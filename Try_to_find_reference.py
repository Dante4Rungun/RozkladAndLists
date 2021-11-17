import xlrd, xlwt
from openpyxl import Workbook
import re
import csv
def recordCsvFile(rozklad):
    with open ('Імпорт розклад.csv','w',newline='',encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file,delimiter=';')
        writer.writerows(rozklad)

def getSheetName(filename):
    pointSheetObj = []
    TeamPointWorkbook = xlrd.open_workbook(filename)
    pointSheets = TeamPointWorkbook.sheet_names()

    for i in pointSheets:
        pointSheetObj.append(tuple((TeamPointWorkbook.sheet_by_name(i),i)))
    result = []
    for i in range(len(pointSheetObj)):
        result.append(pointSheetObj[i][1])
    return result

#-------------------Take rows and cols counts----------------#
def rowsCounter(filename,n,sheetNumber = 0):
    rb = xlrd.open_workbook(filename,formatting_info=True)
    sheet = rb.sheet_by_index(sheetNumber)
    print('SheetNumber = ',sheetNumber)
    count = 0
    for row in range(sheet.nrows):
        if (sheet.row_values(row)[n] != ''):
            count += 1
    return count

def colsCounter(filename):
    rb = xlrd.open_workbook(filename,formatting_info=True)
    sheet = rb.sheet_by_index(0)
    count = 0
    for col in range(sheet.ncols):
        if (sheet.row_values(0)[col] != ''):
            count +=1
    return count
#----------------------Strings methods------------------------#
def firstUpper(mystr):
    split = mystr.split()
    if (len(split) > 1):
        fl = split[0]
        fl = ' ' + fl
        w1 = fl[1]
        w1Upper = w1.upper()
        w1 = ' ' + w1
        fl = re.sub(w1,w1Upper, fl)
        mystr = fl
        for i in range(1,len(split)):
            mystr = mystr + ' ' + split[i]
    return mystr

def checkTitle(string):
    if ("'" in string):
        index = string.index("'")
        string = re.sub(string[index+1],string[index+1].lower(),string)
    return string

def spaceDel(string):
    if(' ' in string):
        string = re.sub(' ','',string)
    return string

def replaceSympols(string):
    if('`' in string):
        string = re.sub('`',"'",string)
    if('᾽' in string):
        string = re.sub('᾽',"'",string)
    if('i' in string):
        string = re.sub('i',"і",string)
    if('0' in string):
        string = re.sub('0','о',string)
    if('c' in string):
        string = re.sub('c','с',string)
    if('o' in string):
        string = re.sub('o','о',string)
    return string

#--------------Student's and teacher's lists------------------#

def teachersList(filename,schoolName):
    if (filename != ''):
        rb = xlrd.open_workbook(filename,formatting_info=True)
        sheet = rb.sheet_by_index(0)
        s = []
        for i in range(0,rowsCounter(filename,0)):
            string = sheet.row_values(i)[0]
            if('`' in string):
                string = re.sub('`',"'",string)
            if('i' in string):
                string = re.sub('i',"і",string)
            PIB = string.split()
            
            s.append([replaceSympols(checkTitle(firstUpper((PIB[0].strip()).title()))),replaceSympols(checkTitle(firstUpper((PIB[1].strip()).title()))),replaceSympols(checkTitle(firstUpper((PIB[2].strip()).title()))),schoolName.strip()])

        wb = Workbook()
        ws = wb.active

        for subarray in s:
            ws.append(subarray)

        wb.save('Імпорт списка вчителів {}.xlsx'.format(schoolName))

    return s

def studentsList(filename,schoolName):
    if (filename != ''):
        rb = xlrd.open_workbook(filename,formatting_info=True)
        sheet = rb.sheet_by_index(0)
        s = []
        for i in range(0,rowsCounter(filename,0)):
            string = sheet.row_values(i)[0]
            if('`' in string):
                string = re.sub('`',"'",string)
            if('i' in string):
                string = re.sub('i',"і",string)
            PIB = string.split()
            clas = str(sheet.row_values(i)[1])
            paralel = str(sheet.row_values(i)[2])
            try:
                s.append([replaceSympols(checkTitle(firstUpper((PIB[0].strip()).title()))),replaceSympols(checkTitle(firstUpper((PIB[1].strip()).title()))),replaceSympols(checkTitle(firstUpper((PIB[2].strip()).title()))),schoolName.strip(),paralel.strip(),clas.strip()])
            except: print('-')
#spaceDel((clas.strip()).upper())
    wb = Workbook()
    ws = wb.active

    for subarray in s:
        ws.append(subarray)

    wb.save('Імпорт списка учнів {}.xlsx'.format(schoolName))
    return s

#-----------------------------After fix methods------------------------------#
def find34 (filename,clas):
    rb = xlrd.open_workbook(filename,formatting_info=True)
    sheet = rb.sheet_by_index(0)

    s = []

    needTime = ['10:20-11:05','11:25-12:10']

    for i in range(0,rowsCounter(filename,1)):
        if ((sheet.row_values(i)[8] in needTime) and (sheet.row_values(i)[3] == clas)):
            subj = sheet.row_values(i)[1].strip()
            teacher = sheet.row_values(i)[2].strip()
            clas = sheet.row_values(i)[3].strip()
            begin = sheet.row_values(i)[4]
            end = sheet.row_values(i)[5]
            day = sheet.row_values(i)[6].strip()
            period= sheet.row_values(i)[7].strip()
            time = sheet.row_values(i)[8]
            
            s.append(['',subj,teacher,clas,begin,end,day,period,time,'09-11-2020'])
    return s

def findNeedSubj(filename):
    clas = ['5-А','5-Б','6-А','6-Б','6-В','7-В','9-Б','10-А','10-Б','11-А']
    result = []
    for i in range(len(clas)):
        result = result + find34(filename,clas[i])
    return result

def changeSpace(mass):
    for i in range(len(mass)):
        for j in range(len(mass[i])):
            if ( ' ' in mass[i][j]):
                mass[i][j] = re.sub(' ','/',mass[i][j])
                #print('+')
    return mass

def uniqNullMass(mass):
    result = []   
    while(len(mass) != 0):    
        for line in mass:
            indexes = []
            n = 0
            result.append([line[0],line[1]])
            for i in range(len(mass)):
                if ((mass[i][0] == line[0]) and (mass[i][1] ==line[1])):
                    indexes.append(i)
            for i in range(len(indexes)):
                if(len(mass) != 0):
                    del mass[indexes[i]-n]
                n += 1
    return result

def rebornTextEditList(fixLine):
    fixData = []
    for i in range(0,len(fixLine)-1,3):
        #print('i = ',i)
        #print('i+2 = ',i+2)
        fixData.append([fixLine[i],fixLine[i+1],fixLine[i+2]])
    return fixData

def changeSlash(mass):
    for i in range(len(mass)):
        for j in range(len(mass[i])):
            if ( '/' in mass[i][j]):
                mass[i][j] = re.sub('/',' ',mass[i][j])
                #print('+')
    return mass

def ResultFix(fix,r):
    for i in range(0,len(r)):
        clas = r[i][3]
        subj = r[i][1]
        for j in range(0,len(fix)):
            if (fix[j][1] == clas and fix[j][0].lower() == subj.lower()):
                r[i][2] = fix[j][2]

    recordCsvFile(r)
    return r
                
#-----------------------Rozklad methods--------------------------------------#
def timeInput(filename):
    rb = xlrd.open_workbook(filename,formatting_info=True)
    sheet = rb.sheet_by_index(0)
    time = []
    timeRows = rowsCounter(filename,0)
    for i in range(0,rowsCounter(filename,0)):
        timeLine = []
        for j in range(0,colsCounter(filename)):
            timeLine.append(sheet.row_values(i)[j])
        time.append(timeLine)
    return time,timeRows

def findClasColumnNumber(filename,clas):
    rb = xlrd.open_workbook(filename,formatting_info=True)
    sheet = rb.sheet_by_index(0)
    columnNumber = 0
    for j in range(0,colsCounter(filename)):
        if (type(sheet.row_values(0)[j]) == 'string'):
            needClas = sheet.row_values(0)[j].strip()
        else: needClas = str(sheet.row_values(0)[j]).strip()     
        if (needClas == clas.strip()):
            columnNumber = j+1
    return columnNumber

def findNeedSheetNameIndex(filename,clas):
    needIndex = ''
    clasList = getSheetName(filename)
    for i in range(len(clasList)):
        if clasList[i] == clas:
            needIndex = i
            break
    return needIndex

def timeDistributionSheets(filename,timeRows,timeCounter,clasList):#вместо двух строк сделать список из n
    timeCounter += 1
    print(timeCounter)
    if (timeRows > 1): #если таймкаунтер  = 1, то идём до первого класа, если 2 то до второго
        timeBarrier = findNeedSheetNameIndex(filename,clasList[timeCounter-1])
        #print(timeBarrier)
    else: timeBarrier = colsCounter(filename)
    return timeBarrier,timeCounter           

def timeDistribution(filename,timeRows,timeCounter,clasList):#вместо двух строк сделать список из n
    timeCounter += 1
    if (timeRows > 1): #если таймкаунтер  = 1, то идём до первого класа, если 2 то до второго
        timeBarrier = findClasColumnNumber(filename,clasList[timeCounter-1])
        #print(timeBarrier)
    else: timeBarrier = colsCounter(filename)
    return timeBarrier,timeCounter
    
def takeAfter (filename):
    rb = xlrd.open_workbook(filename,formatting_info=True)
    sheet = rb.sheet_by_index(0)
    rozAfter = []
    for i in range(0,rowsCounter(filename,1)):
        rozAfter.append([sheet.row_values(i)[0],sheet.row_values(i)[1],sheet.row_values(i)[2],sheet.row_values(i)[3],sheet.row_values(i)[4],sheet.row_values(i)[5],sheet.row_values(i)[6],sheet.row_values(i)[7],sheet.row_values(i)[8]])

    return rozAfter

def findNull(mass):
    count = 0
    nullMass = []
    for i in range(len(mass)):
        if ( mass[i][2] == ''):
            count +=1
            nullMass.append([mass[i][1],mass[i][3]])
                   
    return nullMass

def TeachSubj(FileName):
    rb = xlrd.open_workbook(FileName,formatting_info=True)
    sheet = rb.sheet_by_index(0)
    z = []

    for i in range(0,rowsCounter(FileName,0)):
        if (type(sheet.row_values(i)[0]) == 'string'):
            clas = sheet.row_values(i)[0].strip()
        else: clas = str(sheet.row_values(i)[0]).strip()
        z.append([clas,sheet.row_values(i)[1].strip(),(checkTitle(replaceSympols(sheet.row_values(i)[2].strip()).title()))])

    return z

def Rozklad(filename,sheetNumber,filenameRings,firstData,secondData,clasList):
    rb = xlrd.open_workbook(filename,formatting_info=True)
    sheet = rb.sheet_by_index(sheetNumber)
    #time = [['13:30-13:50','14:00-14:20','14:40-15:00','15:10-15:30','15:50-16:10','16:20-16:40','17:00-17:20','17:30-17:50'],
            #['09:00-09:30','09:40-10:10','10:30-11:00','11:10-11:40','12:00-12:30','12:40-13:10','13:30-14:00','14:10-14:40']]
    time,timeRows = timeInput(filenameRings)
    date = []
    date.append(firstData)
    date.append(secondData)
    s = []
    s.append(['Аудиторія','Предмет','Вчитель','Клас','Початок семестру','Кінець семестру','День тижня','Періодичність','Час уроку'])
    timeCounter = 0
    #timeBarrier = 2
    for j in range(2,colsCounter(filename)):
        if (j == 2 or j == timeBarrier):
            timeBarrier,timeCounter = timeDistribution(filename,timeRows,timeCounter,clasList)
        for i in range(1,rowsCounter(filename,0)):
            if (sheet.row_values(i)[j] == ''):
                continue
            else:
                if (j < timeBarrier):#создать список значений от 1-го до 4-ех и проверять с его помощью
                    n = timeCounter - 1
                subj = firstUpper(sheet.row_values(i)[j].strip())
                day = sheet.row_values(i)[0].strip()
                if (type(sheet.row_values(0)[j]) == 'string'):
                    clas = sheet.row_values(0)[j].strip()
                else: clas = str(sheet.row_values(0)[j]).strip()
                numb = sheet.row_values(i)[1]
                number = int(numb)
                teacher = ''
                if ('/' in subj):
                    twoSubj = subj.split('/')
                    s.append(['',firstUpper(twoSubj[0].strip()),teacher,spaceDel(clas),spaceDel(date[0].strip()),spaceDel(date[1].strip()),day,'Щотижня',(time[n][number-1]).strip(),'+'])
                    s.append(['',firstUpper(twoSubj[1].strip()),teacher,spaceDel(clas),spaceDel(date[0].strip()),spaceDel(date[1].strip()),day,'Щотижня',(time[n][number-1]).strip(),'+'])
                elif('-' in subj):
                    twoSubj = subj.split('-')
                    s.append(['',firstUpper(twoSubj[0].strip()),teacher,spaceDel(clas),spaceDel(date[0].strip()),spaceDel(date[1].strip()),day,'Щотижня',(time[n][number-1]).strip(),'+'])
                    s.append(['',firstUpper(twoSubj[1].strip()),teacher,spaceDel(clas),spaceDel(date[0].strip()),spaceDel(date[1].strip()),day,'Щотижня',(time[n][number-1]).strip(),'+'])
                else:
                    s.append(['',subj,teacher,clas,spaceDel(date[0].strip()),spaceDel(date[1].strip()),day,'Щотижня',time[n][number-1]])
    return s

def Result(t,r):
    for i in range(0,len(r)):
        clas = r[i][3]
        subj = r[i][1]
        for j in range(0,len(t)):
            if (t[j][0] == clas and t[j][1].lower() == subj.lower()):
                r[i][2] = t[j][2]
    return r

def fewSheetsResult(filename1,filename2,sheetsString,filenameRings,firstData,secondData,clasLists,schoolName): 
    sheetList = sheetsString.split()
    clasList = clasLists.split()
    for i in range(0,len(sheetList)):
        t = TeachSubj(filename2)  #('ІЗОШ №6 Список вчителів і предметів (1).xls')
        r = Rozklad(filename1,int(sheetList[i]),filenameRings,firstData,secondData,clasList) #стандарт
        rez = Result(t,r)
        #findNull(rez)
        
        #for line in rez:
         #   print(line)
        wb = Workbook()
        ws = wb.active

        for subarray in rez:
            ws.append(subarray)

        wb.save('Імпорт розклад {0} тиждень {1}.xlsx'.format(i+1,schoolName))
        #print('-----------------------------------------------------------------------------------------------------')
        return rez
#-------------------record in file def-------------------#


    #C:\Users\Vekto_000\Desktop\TestInterface\Імпорт розклад.csv










    
